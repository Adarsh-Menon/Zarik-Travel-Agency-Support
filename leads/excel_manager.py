import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import LEADS_EXCEL_PATH

COLUMNS = [
    "lead_id", "name", "telegram_handle", "telegram_id", "phone",
    "destination", "travel_dates", "budget", "group_size", "interests",
    "status", "itinerary_summary", "created_at", "last_updated", "notes"
]

HEADER_FILL = PatternFill("solid", fgColor="1B4F72")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
STATUS_COLORS = {
    "New": PatternFill("solid", fgColor="AED6F1"),
    "Contacted": PatternFill("solid", fgColor="F9E79F"),
    "Converted": PatternFill("solid", fgColor="ABEBC6"),
    "Lost": PatternFill("solid", fgColor="F5B7B1"),
}
COL_WIDTHS = {
    "A": 12, "B": 18, "C": 18, "D": 14, "E": 16,
    "F": 18, "G": 16, "H": 14, "I": 12, "J": 25,
    "K": 14, "L": 35, "M": 20, "N": 20, "O": 30
}
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def _ensure_workbook():
    os.makedirs(os.path.dirname(LEADS_EXCEL_PATH), exist_ok=True)
    if not os.path.exists(LEADS_EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        for col_letter, width in COL_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
        ws.auto_filter.ref = f"A1:O1"
        ws.freeze_panes = "A2"
        wb.save(LEADS_EXCEL_PATH)
    return load_workbook(LEADS_EXCEL_PATH)


def _next_lead_id(ws) -> str:
    max_num = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and str(row[0]).startswith("ZRK-"):
            try:
                num = int(str(row[0]).split("-")[1])
                max_num = max(max_num, num)
            except ValueError:
                pass
    return f"ZRK-{max_num + 1:03d}"


def add_lead(
    name: str, telegram_handle: str, telegram_id: int,
    destination: str = "", travel_dates: str = "", budget: str = "",
    group_size: int = 0, interests: str = "", itinerary_summary: str = "",
    phone: str = "", notes: str = ""
) -> str:
    wb = _ensure_workbook()
    ws = wb["Leads"]
    lead_id = _next_lead_id(ws)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    row_data = [
        lead_id, name, telegram_handle, telegram_id, phone,
        destination, travel_dates, budget, group_size, interests,
        "New", itinerary_summary, now, now, notes
    ]
    row_num = ws.max_row + 1
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = Font(name="Arial", size=10)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=col_idx in [10, 12, 15])

    # Apply status color
    status_cell = ws.cell(row=row_num, column=11)
    status_cell.fill = STATUS_COLORS.get("New", PatternFill())

    wb.save(LEADS_EXCEL_PATH)
    return lead_id


def update_lead(lead_id: str, **fields):
    wb = _ensure_workbook()
    ws = wb["Leads"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == lead_id:
            for field, value in fields.items():
                if field in COLUMNS:
                    col_idx = COLUMNS.index(field) + 1
                    row[col_idx - 1].value = value
                    if field == "status" and value in STATUS_COLORS:
                        row[col_idx - 1].fill = STATUS_COLORS[value]
            # Update last_updated
            row[COLUMNS.index("last_updated")].value = datetime.now().strftime("%Y-%m-%d %H:%M")
            wb.save(LEADS_EXCEL_PATH)
            return True
    return False


def get_lead(lead_id: str) -> dict | None:
    wb = _ensure_workbook()
    ws = wb["Leads"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == lead_id:
            return dict(zip(COLUMNS, row))
    return None


def find_lead_by_telegram_id(telegram_id: int) -> dict | None:
    wb = _ensure_workbook()
    ws = wb["Leads"]
    found = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] == telegram_id:
            found = dict(zip(COLUMNS, row))  # return latest
    return found


def get_all_leads(status_filter: str = None) -> list[dict]:
    wb = _ensure_workbook()
    ws = wb["Leads"]
    leads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        lead = dict(zip(COLUMNS, row))
        if status_filter and lead.get("status") != status_filter:
            continue
        leads.append(lead)
    return leads


def get_lead_stats() -> dict:
    leads = get_all_leads()
    stats = {"total": len(leads), "by_status": {}}
    for lead in leads:
        s = lead.get("status", "Unknown")
        stats["by_status"][s] = stats["by_status"].get(s, 0) + 1
    return stats
